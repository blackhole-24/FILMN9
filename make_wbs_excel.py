# -*- coding: utf-8 -*-
"""FILMN9 WBS Excel 생성 — 캡쳐 형식 그대로"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "FILMN9 WBS"

# ── 색상 정의 ─────────────────────────────────────
HEADER_BG = "1F4E3D"   # 진한 초록
SPRINT_BG = "B8E0C9"   # 연한 초록 (스프린트 헤더)
KICKOFF_BG = "B8E0C9"
# 담당자
A_HJ  = "2C3E50"   # 이형주 — 짙은 회색
A_HW  = "C0392B"   # 고휘주 — 빨강
A_YT  = "27AE60"   # 유태상 — 초록
A_SS  = "5DADE2"   # 손소희 — 하늘
A_ALL = "8E44AD"   # 전원 — 보라
# 상태
S_OK   = "C8E6C9"  # 완료 — 초록
S_DOING = "FFF59D" # 진행 중 — 노랑
S_TODO = "BBDEFB"  # 할 일 — 하늘
S_CHK  = "FFCC80"  # 확인 필요 — 주황
S_REJ  = "FFCDD2"  # 기각 — 빨강
S_HOLD = "E1BEE7"  # 보류 — 보라

WHITE = "FFFFFF"
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fill(color): return PatternFill("solid", fgColor=color)

def assignee_color(name):
    """담당자별 배경색"""
    if not name:
        return None
    name = name.strip()
    if "전원" in name: return A_ALL
    if "이형주" in name: return A_HJ
    if "고휘주" in name: return A_HW
    if "유태상" in name: return A_YT
    if "손소희" in name: return A_SS
    return None

def status_color(status):
    if not status: return None
    if "완료" in status: return S_OK
    if "진행" in status: return S_DOING
    if "할 일" in status or "할일" in status: return S_TODO
    if "확인" in status: return S_CHK
    if "기각" in status: return S_REJ
    if "보류" in status: return S_HOLD
    return None

# ── 헤더 ───────────────────────────────────────────
ws["A1"] = "FILMN9 WBS"
ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
ws["A1"].fill = fill(HEADER_BG)
ws.merge_cells("A1:H1")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 28

# 컬럼 헤더 (캡쳐 그대로)
headers = ["Group", "Task", "Subtask", "Link", "Assignee", "Start Date", "End Date", "Period", "Status"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=i, value=h)
    c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    c.fill = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
ws.row_dimensions[2].height = 30

# 컬럼 폭
widths = [12, 22, 60, 12, 14, 12, 12, 8, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── 데이터 ─────────────────────────────────────────
# 형식: (Group, Task, Subtask, Link, Assignee, Start, End, Period, Status, is_sprint_header)
data = [
    # ─────── 킥오프 ───────
    ("킥오프", "", "", "", "", "04/22", "04/28", 1, "완료", True),
    ("04/22\n~\n04/24", "킥오프\n(프로젝트 헌정 작성)", "팀명 / 역할 / 그라운드룰 확정", "[팀빌딩]", "전원", "04/22", "04/28", 1, "완료", False),
    ("", "", "주제선정", "[드라이브]", "전원", "04/22", "04/23", 1, "완료", False),
    ("", "", "스프린트 계획서 확정", "[드라이브]", "전원", "04/22", "04/24", 1, "완료", False),
    ("", "", "프로젝트 헌장서 작성", "[드라이브]", "전원", "04/24", "04/24", 1, "완료", False),
    ("", "", "서비스 범위 + 페르소나 설정", "[드라이브]", "손소희", "04/27", "04/27", 1, "완료", False),
    ("", "", "프로젝트 WBS 작성", "[드라이브]", "전원", "04/27", "04/27", 1, "완료", False),

    # ─────── 스프린트 1 ───────
    ("스프린트 1", "", "", "", "", "04/27", "04/28", 2, "완료", True),
    ("04/27\n~\n04/28", "기획", "기획서 작성", "[프로젝트 기획서]", "전원", "04/27", "04/28", 1, "완료", False),
    ("", "", "프로젝트 WBS 작성", "", "전원", "04/27", "04/28", 2, "완료", False),
    ("", "", "프로젝트 관리 도구 (노션)", "[노션]", "이형주", "04/27", "04/28", 2, "완료", False),
    ("", "", "개발 소스 저장소 (깃) 생성", "[github]", "이형주", "04/27", "04/28", 2, "완료", False),
    ("", "", "산출물 저장소", "[드라이브]", "이형주", "04/27", "04/28", 2, "완료", False),
    ("", "", "서비스 범위 + 페르소나 확정", "[드라이브]", "손소희", "04/27", "04/28", 2, "완료", False),
    ("", "", "요구사항 정의서", "[드라이브]", "손소희", "04/27", "04/28", 2, "완료", False),
    ("", "", "스프린트 1 리뷰", "[드라이브]", "전원", "04/27", "04/28", 2, "완료", False),

    # ─────── 스프린트 2 ───────
    ("스프린트 2", "", "", "", "", "04/29", "05/09", 11, "진행 중", True),
    ("04/29\n~\n05/09", "개발 환경 구축", "Python 가상환경·VS Code·패키지 설치", "[드라이브]", "전원", "04/30", "04/30", 1, "완료", False),
    ("", "", "MySQL DB 생성 + 기본 테이블 DDL", "[드라이브]", "전원", "05/08", "05/08", 9, "진행 중", False),
    ("", "데이터 수집", "기업 재무제표(2025년 사업보고서) 일괄 수집", "[드라이브]", "전원", "05/08", "05/08", 9, "완료", False),
    ("", "", "ECOS API: Rf 3년국고채 자동 수집", "[드라이브]", "고휘주", "05/08", "05/08", 9, "완료", False),
    ("", "", "KOFIA BOND: Kd 신용등급별 채권수익률 CSV 수집", "[드라이브]", "고휘주", "05/08", "05/08", 9, "확인 필요", False),
    ("", "", "KOFIA BOND: Kd → MySQL 업로드 스크립트", "[드라이브]", "고휘주", "05/08", "05/08", 9, "완료", False),
    ("", "", "KRX: 평가기준일 시가총액 수집 (주가×발행주식수)", "[드라이브]", "유태상", "05/08", "05/08", 9, "진행 중", False),
    ("", "", "종목별 주가·지수 수집 (Beta 계산 기초 데이터)", "[드라이브]", "유태상", "05/08", "05/08", 9, "기각", False),
    ("", "", "유사기업 산정 알고리즘 구축", "[드라이브]", "유태상", "05/08", "05/08", 9, "진행 중", False),
    ("", "밸류에이션\n계산 구현·검증", "WACC Python 계산 로직 구현", "[드라이브]", "유태상", "05/08", "05/08", 3, "진행 중", False),
    ("", "", "FCF 계산 Python 구현 (EBIT×(1-Tc)+D&A-CAPEX-ΔNWC)", "[드라이브]", "전원", "05/08", "05/08", 3, "진행 중", False),
    ("", "", "DCF 3-Way 시나리오 계산 구현", "[드라이브]", "전원", "05/08", "05/08", 3, "진행 중", False),
    ("", "", "멀티플 계산 구현 (EV/EBITDA·P/E·EV/Revenue)", "[드라이브]", "전원", "05/08", "05/08", 3, "진행 중", False),
    ("", "", "테스트 기업 3~5개 적용·결과 검증", "[드라이브]", "전원", "05/08", "05/08", 3, "진행 중", False),
    ("", "와이어프레임", "Figma: WACC 입력폼·결과카드·시나리오 히트맵 화면", "[드라이브]", "손소희", "05/08", "05/08", 3, "진행 중", False),
    ("", "", "Figma: DCF Bear/Base/Bull 결과·민감도 5×5 표 화면", "[드라이브]", "손소희", "05/08", "05/08", 3, "진행 중", False),
    ("", "", "Figma: 멀티플 Peer 비교표·적정주가 밴드 화면", "[드라이브]", "손소희", "05/08", "05/08", 3, "진행 중", False),
    ("", "", "Figma: 일반투자자용 간략 뷰", "[드라이브]", "손소희", "05/08", "05/08", 3, "진행 중", False),
    ("", "Sprint 2 마무리", "MVP 범위 최종 확정 (WACC+DCF+멀티플 통합)", "[드라이브]", "전원", "05/08", "05/08", 2, "할 일", False),
    ("", "", "Sprint 2 리뷰 발표", "[드라이브]", "전원", "05/08", "05/09", 1, "할 일", False),
    ("", "", "1차 멘토링 준비·참석 (5/9)", "[드라이브]", "전원", "05/09", "05/09", 1, "할 일", False),

    # ─────── 스프린트 3 (1차 배포) ───────
    ("스프린트 3", "", "1차 배포 (데모 시연)", "", "", "05/12", "05/22", 11, "완료", True),
    ("05/12\n~\n05/22\n(5/26 실배포)", "[기업개요]\n기획·설계", "DartPoint AI 벤치마크 분석 + 기업개요 명세서", "[드라이브]", "이형주\n고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "", "기업개요 탭 HTML 와이어프레임 (DartPoint 구조)", "[드라이브]", "이형주\n고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "", "AI 기업 히스토리 브리핑 기획 (n장 포맷)", "[드라이브]", "이형주\n고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "[기업개요]\n데이터 수집·파이프라인", "DART API 기업 기본정보 (대표·업종·주소·주주)", "[드라이브]", "이형주\n고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "", "네이버 뉴스 RSS 연동 + 종목별 뉴스 파이프라인", "[드라이브]", "이형주\n고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "", "yfinance 실시간 주가·시총 (15분 지연, 60초 폴링)", "[드라이브]", "이형주\n고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "", "KRX OHLCV 일봉 차트 데이터 (1Y 기본)", "[드라이브]", "이형주\n고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "[기업개요]\nAI 히스토리 브리핑", "DART 사업보고서 RAG 청크 (4팀 분담 2,596개)", "[드라이브]", "고휘주\n전원", "05/12", "05/22", 11, "완료", False),
    ("", "", "gpt-5-mini 자동 요약 (사업개요·고객·리스크·근거)", "[드라이브]", "고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "", "히스토리 브리핑 142개 종목 MongoDB 적재", "[드라이브]", "고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "", "신뢰도 등급(high/medium/low) 자동 분류", "[드라이브]", "고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "", "★ 데모 3종목 (아모레·삼성전기·NAVER) 브리핑", "[드라이브]", "이형주\n고휘주", "05/12", "05/22", 11, "완료", False),
    ("", "Valuation\n[기업가치]\nLLM 파이프라인", "LangChain LCEL 체인 구조 (gpt-5-mini)", "[드라이브]", "유태상\n손소희", "05/12", "05/22", 11, "완료", False),
    ("", "", "IBD 분류 파이프라인 (LLM → 계정명 포함/제외) Python 합산", "[드라이브]", "유태상\n손소희", "05/12", "05/22", 11, "완료", False),
    ("", "", "GICS 업종 자동 분류 (DART 텍스트 → LLM → GICS 코드)", "[드라이브]", "유태상\n손소희", "05/12", "05/22", 11, "완료", False),
    ("", "Valuation\nBeta 계산", "KRX 회귀분석 → β_L 산출 (5년월·3년월·2년주)", "[드라이브]", "유태상\n손소희", "05/12", "05/22", 11, "완료", False),
    ("", "", "Hamada 방정식 → β_U 언레버링 (β_U = β_L ÷ (1+(1-Tc)×D/E))", "[드라이브]", "유태상\n손소희", "05/12", "05/22", 11, "완료", False),
    ("", "", "Damodaran 업종평균 β_U 교차검증 (Gap > 0.3 σ 경고)", "[드라이브]", "유태상\n손소희", "05/12", "05/22", 11, "완료", False),
    ("", "Valuation\nDCF 고도화", "Damodaran 섹터 평균 매출성장률 (Base g 자동화)", "[드라이브]", "유태상\n손소희", "05/12", "05/22", 11, "완료", False),
    ("", "", "과거 3년 FCFF CAGR 자동 계산 (Bull g 자동화)", "[드라이브]", "유태상\n손소희", "05/12", "05/22", 11, "완료", False),
    ("", "", "WACC × g 민감도 히트맵 (5×5 = 25개 시나리오, Plotly)", "[드라이브]", "유태상\n손소희", "05/12", "05/22", 11, "완료", False),
    ("", "UI 구현", "Next.js 16 + React 19 환경 + v0.dev UI 생성", "[드라이브]", "이형주", "05/19", "05/22", 3, "완료", False),
    ("", "", "종목 페이지 (3개 탭: 기업개요·밸류·AI) 1700줄 구현", "[드라이브]", "이형주", "05/19", "05/22", 3, "완료", False),
    ("", "", "백엔드 FastAPI 15+ 엔드포인트 (/api/overview 등)", "[드라이브]", "이형주", "05/19", "05/22", 3, "완료", False),
    ("", "인프라", "SQLite DB 스키마 (9 테이블) + ERD 작성", "[드라이브]", "이형주", "05/19", "05/22", 3, "완료", False),
    ("", "", "MongoDB Atlas 연결 + 히스토리 폴백 모드", "[드라이브]", "이형주\n고휘주", "05/19", "05/22", 3, "완료", False),
    ("", "1차 배포\nSprint 3 마무리", "start.bat 원클릭 기동 + 로컬호스트 배포", "[드라이브]", "이형주", "05/20", "05/22", 3, "완료", False),
    ("", "", "Sprint 3 리뷰 발표 (1차 배포 데모)", "[드라이브]", "전원", "05/26", "05/26", 1, "완료", False),
    ("", "3차 멘토링", "3차 멘토링 준비·참석", "[드라이브]", "전원", "05/22", "05/22", 1, "완료", False),

    # ─────── 스프린트 4 (2차 배포) ───────
    ("스프린트 4", "", "2차 배포 (산출물 통합 + 고도화)", "", "", "05/26", "06/05", 11, "진행 중", True),
    ("05/26\n~\n06/05", "[기업개요]\n재무정보 정확성", "사업보고서 XML 1,295 → 2,844개 확장 다운로드", "[드라이브]", "이형주", "05/26", "05/28", 3, "완료", False),
    ("", "", "XML inline XBRL 파싱 (FY2025·2024·2023 3개년)", "[드라이브]", "이형주", "05/26", "05/28", 3, "완료", False),
    ("", "", "DB 적재 1,861개 종목 · 625,218건", "[드라이브]", "이형주", "05/26", "05/28", 3, "완료", False),
    ("", "", "백엔드 API: 연결/별도 토글 + 음수 처리 + 우선정렬", "[드라이브]", "이형주", "05/27", "05/28", 2, "완료", False),
    ("", "", "NAVER 등 외감법인 DART API 보충", "[드라이브]", "이형주", "05/27", "05/28", 2, "완료", False),
    ("", "[기업개요]\nUX 개선", "기업개요·경영인 2-col 레이아웃", "[드라이브]", "이형주", "05/26", "05/30", 5, "완료", False),
    ("", "", "주가차트 1Y 기본 + 줌인·페이지스크롤 분리", "[드라이브]", "이형주", "05/26", "05/30", 5, "완료", False),
    ("", "", "헤더: 실시간 주가 + (Yahoo Finance KST · ~15분 지연)", "[드라이브]", "이형주", "05/26", "05/30", 5, "완료", False),
    ("", "", "업종(WICS) 헤더 추가 + 시총 자동 계산", "[드라이브]", "이형주", "05/26", "05/30", 5, "완료", False),
    ("", "", "히스토리 브리핑 + 최신 뉴스 섹션 통합", "[드라이브]", "이형주\n고휘주", "05/26", "05/30", 5, "완료", False),
    ("", "", "기업건전성 (부채비율·유동비율·영업이익률) 자동계산", "[드라이브]", "이형주", "05/26", "05/30", 5, "완료", False),
    ("", "[기업개요]\nAI 브리핑 확장", "briefs_final 2,526개 MongoDB 일괄 적재", "[드라이브]", "고휘주", "05/26", "05/27", 2, "완료", False),
    ("", "", "차별성 ①: 감사의견 자동 추출 (DART 감사보고서)", "[드라이브]", "고휘주", "05/26", "06/05", 11, "할 일", False),
    ("", "", "차별성 ②: 리스크 요소 자동 감지 (소송·신용등급·CB)", "[드라이브]", "고휘주", "05/26", "06/05", 11, "할 일", False),
    ("", "", "차별성 ③: 단일판매 공급계약 공시 자동 추출", "[드라이브]", "고휘주", "05/26", "06/05", 11, "할 일", False),
    ("", "", "차별성 ④: 주가 영향 주요 요인 AI 분석", "[드라이브]", "고휘주", "05/26", "06/05", 11, "할 일", False),
    ("", "Valuation\n1차 산출물 통합", "태상님 ui_payload JSON (3종목) 백엔드 통합", "[드라이브]", "유태상\n이형주", "05/26", "05/30", 5, "할 일", False),
    ("", "", "Bear/Base/Bull 카드 실데이터 연결 (적정가 검증)", "[드라이브]", "유태상\n이형주", "05/26", "05/30", 5, "할 일", False),
    ("", "", "DCF·WACC·멀티플·민감도·토네이도·피어 비교 6탭 통합", "[드라이브]", "유태상\n이형주", "05/26", "05/30", 5, "할 일", False),
    ("", "", "가치진단 5종 (EPV·성장 프리미엄·시장 함의 성장률 등)", "[드라이브]", "유태상", "05/26", "06/05", 11, "할 일", False),
    ("", "", "삼성전기 등 비정상 수치 재검증 (PER 268배 등)", "[드라이브]", "유태상", "05/26", "06/05", 11, "할 일", False),
    ("", "데이터 보충", "외감법인 529개 JSONL/임베딩 활용 보충", "[드라이브]", "이형주\n고휘주", "06/02", "06/05", 4, "할 일", False),
    ("", "", "OHLCV 자동 갱신 스케줄러 (매일 장 마감 후)", "[드라이브]", "이형주", "06/02", "06/05", 4, "할 일", False),
    ("", "인프라 통합", "GitHub 개발환경_통합/ + 인수인계서 배포", "[드라이브]", "이형주", "05/27", "05/28", 2, "완료", False),
    ("", "", "data 패키지 8개 zip Google Drive 공유", "[드라이브]", "이형주", "05/27", "05/28", 2, "완료", False),
    ("", "", "팀원 PC 환경 통합 (개인 PC + KPMG PC 동기화)", "[드라이브]", "전원", "05/27", "06/05", 10, "진행 중", False),
    ("", "2차 배포\nSprint 4 마무리", "2차 배포 (산출물 통합본)", "[드라이브]", "전원", "06/04", "06/05", 2, "할 일", False),
    ("", "", "Sprint 4 리뷰 발표", "[드라이브]", "전원", "06/05", "06/05", 1, "할 일", False),
    ("", "4차 멘토링", "4차 멘토링 준비·참석", "[드라이브]", "전원", "06/05", "06/05", 1, "할 일", False),

    # ─────── 스프린트 5 (3차 배포 + 최종) ───────
    ("스프린트 5", "", "3차 배포 + 최종 완성", "", "", "06/08", "06/19", 12, "할 일", True),
    ("06/08\n~\n06/19", "[기업개요]\n완성도 강화", "4가지 차별성 통합 (AI 종합 분석 보고서)", "[드라이브]", "고휘주", "06/08", "06/19", 12, "할 일", False),
    ("", "", "Sankey 손익 흐름도 3종목 재렌더링 + 자동화", "[드라이브]", "이형주\n고휘주", "06/08", "06/15", 8, "할 일", False),
    ("", "", "신용등급 추이 차트 (2020~2025) 완성", "[드라이브]", "이형주", "06/08", "06/12", 5, "할 일", False),
    ("", "", "재무 하이라이트 위젯 (YoY 증감률 색상 코딩)", "[드라이브]", "이형주", "06/08", "06/12", 5, "할 일", False),
    ("", "[기업개요]\nAI 분석 (Tab3)", "AI 종합분석 SWOT 자동 생성 (실데이터 기반)", "[드라이브]", "고휘주", "06/08", "06/15", 8, "할 일", False),
    ("", "", "RAG 챗봇 정식 엔진 (Chroma 임베딩 + LLM)", "[드라이브]", "이형주\n고휘주", "06/08", "06/15", 8, "할 일", False),
    ("", "", "종목별 추천 질문 자동 생성 (기존 사전매핑 → 동적)", "[드라이브]", "고휘주", "06/08", "06/15", 8, "할 일", False),
    ("", "", "사업보고서 출처 표기 + 신뢰도 등급 표시", "[드라이브]", "고휘주", "06/08", "06/15", 8, "할 일", False),
    ("", "Valuation\n2차 완성", "4-Way 주식 내재가치 비교 바차트 (DCF·PER·PBR·EV/EBITDA)", "[드라이브]", "유태상", "06/08", "06/15", 8, "할 일", False),
    ("", "", "AI 진단·해설 자동 생성 (LLM 평가/저평가 판단)", "[드라이브]", "유태상", "06/08", "06/15", 8, "할 일", False),
    ("", "", "Audit Trail (출처 URL·수집 날짜·계산 공식 전체 기록)", "[드라이브]", "유태상", "06/08", "06/15", 8, "할 일", False),
    ("", "", "Excel·PDF 리포트 내보내기 기능", "[드라이브]", "유태상", "06/12", "06/17", 6, "할 일", False),
    ("", "데이터 정확성", "XML+JSONL+임베딩 결합 — 외감법인 포함 전체 종목", "[드라이브]", "이형주\n고휘주", "06/08", "06/15", 8, "할 일", False),
    ("", "", "데이터 검증 자동화 (사업보고서 vs DB 일치 체크)", "[드라이브]", "이형주", "06/08", "06/12", 5, "할 일", False),
    ("", "", "신규 상장사 자동 추가 파이프라인", "[드라이브]", "이형주", "06/12", "06/17", 6, "할 일", False),
    ("", "인프라 고도화", "Fast API 백엔드 캐싱 (Redis 또는 SQLite 캐시)", "[드라이브]", "이형주", "06/08", "06/15", 8, "할 일", False),
    ("", "", "API 응답 시간 최적화 (1초 이내 목표)", "[드라이브]", "이형주", "06/08", "06/15", 8, "할 일", False),
    ("", "", "모바일 반응형 최적화", "[드라이브]", "이형주", "06/15", "06/19", 5, "할 일", False),
    ("", "(선택)\n클라우드 배포", "AWS EC2 / Vercel 배포 검토·진행", "[드라이브]", "이형주", "06/15", "06/19", 5, "할 일", False),
    ("", "", "도메인·SSL 설정", "[드라이브]", "이형주", "06/15", "06/19", 5, "할 일", False),
    ("", "3차 배포\nSprint 5 마무리", "3차 배포 (최종 완성본)", "[드라이브]", "전원", "06/17", "06/19", 3, "할 일", False),
    ("", "", "Sprint 5 리뷰 발표 (최종 데모)", "[드라이브]", "전원", "06/19", "06/19", 1, "할 일", False),
    ("", "5차 멘토링·발표 준비", "5차 멘토링 준비·참석", "[드라이브]", "전원", "06/16", "06/16", 1, "할 일", False),
    ("", "", "최종 발표 자료 작성 (PPT·시연 영상·문서)", "[드라이브]", "전원", "06/16", "06/19", 4, "할 일", False),
    ("", "", "최종 발표 리허설", "[드라이브]", "전원", "06/19", "06/19", 1, "할 일", False),
    ("", "🎯 최종 발표", "🎯 최종 발표 (FILMN9 데모)", "[드라이브]", "전원", "06/20", "06/20", 1, "할 일", False),
]

# ── 데이터 입력 ────────────────────────────────────
row = 3
for d in data:
    group, task, subtask, link, assignee, start, end, period, status, is_header = d
    if is_header:
        # 스프린트 헤더 (한 줄 강조)
        for c in range(1, 10):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill(SPRINT_BG)
            cell.font = Font(name="맑은 고딕", size=11, bold=True, color="1F4E3D")
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.cell(row=row, column=1, value=group)
        ws.cell(row=row, column=3, value=subtask)
        ws.cell(row=row, column=6, value=start)
        ws.cell(row=row, column=7, value=end)
        ws.cell(row=row, column=8, value=period)
        ws.cell(row=row, column=9, value=status)
        ws.row_dimensions[row].height = 22
    else:
        # 일반 행
        ws.cell(row=row, column=1, value=group)
        ws.cell(row=row, column=2, value=task)
        ws.cell(row=row, column=3, value=subtask)
        ws.cell(row=row, column=4, value=link)
        ws.cell(row=row, column=5, value=assignee)
        ws.cell(row=row, column=6, value=start)
        ws.cell(row=row, column=7, value=end)
        ws.cell(row=row, column=8, value=period)
        ws.cell(row=row, column=9, value=status)

        for c in range(1, 10):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if c == 2 or c == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

        # Group 셀 강조
        if group:
            ws.cell(row=row, column=1).font = Font(name="맑은 고딕", size=9, bold=True)
            ws.cell(row=row, column=1).fill = fill("F5F5F5")

        # Link 파란색
        if link:
            ws.cell(row=row, column=4).font = Font(name="맑은 고딕", size=9, color="1F4E3D", underline="single")

        # Assignee 배경색
        a_color = assignee_color(assignee)
        if a_color:
            ws.cell(row=row, column=5).fill = fill(a_color)
            ws.cell(row=row, column=5).font = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")

        # Start/End 노란 배경
        ws.cell(row=row, column=6).fill = fill("FFF9C4")
        ws.cell(row=row, column=7).fill = fill("FFF9C4")

        # Status 배경
        s_color = status_color(status)
        if s_color:
            ws.cell(row=row, column=9).fill = fill(s_color)
            ws.cell(row=row, column=9).font = Font(name="맑은 고딕", size=9, bold=True)

        ws.row_dimensions[row].height = 28
    row += 1

# 윗 행 고정
ws.freeze_panes = "A3"

# ── 저장 ─────────────────────────────────────────
output = r"C:\Users\Admin\FILMN9\FILMN9_WBS_업데이트_v4.xlsx"
wb.save(output)
print(f"저장 완료: {output}")
print(f"총 {row-3} 행 작성")
