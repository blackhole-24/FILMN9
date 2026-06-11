# -*- coding: utf-8 -*-
"""analyst_report_targets.csv -> xlsx(다운로드 추적용) + txt(체크리스트) 생성"""
import csv, os

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "analyst_report_targets.csv")

rows = []
with open(SRC, "r", encoding="utf-8-sig", newline="") as f:
    reader = csv.reader(f)
    header = next(reader)
    for r in reader:
        if not r or not r[0].strip():
            continue
        rows.append(r)

print(f"읽은 데이터 행수: {len(rows)}")
print(f"헤더: {header}")

# ---------- 1) XLSX 생성 ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "애널리스트리포트_타겟"

hdr_fill = PatternFill("solid", fgColor="1E40AF")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
core_fill = PatternFill("solid", fgColor="DBEAFE")   # 코어 연파랑
ext_fill  = PatternFill("solid", fgColor="FEF3C7")   # 확장 연노랑
thin = Side(style="thin", color="E2E8F0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 헤더 작성
for c, name in enumerate(header, start=1):
    cell = ws.cell(row=1, column=c, value=name)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# 데이터 작성 (종목코드는 텍스트로 보존)
for i, r in enumerate(rows, start=2):
    tier = r[5] if len(r) > 5 else ""
    for c, val in enumerate(r, start=1):
        cell = ws.cell(row=i, column=c, value=val)
        cell.border = border
        if c == 2:  # 종목코드 텍스트
            cell.number_format = "@"
        cell.fill = core_fill if "코어" in tier else ext_fill

# 컬럼 폭
widths = [8, 12, 22, 16, 14, 14, 40, 16]
for c, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A2"           # 헤더 고정
ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"  # 필터

XLSX = os.path.join(BASE, "analyst_report_targets.xlsx")
wb.save(XLSX)
print(f"[OK] 엑셀 저장: {XLSX}")

# ---------- 2) TXT 체크리스트 생성 ----------
TXT = os.path.join(BASE, "analyst_report_targets_LIST.txt")
core = [r for r in rows if len(r) > 5 and "코어" in r[5]]
ext  = [r for r in rows if len(r) > 5 and "확장" in r[5]]

def fmt(r):
    pr, code, name, mkt, cap = r[0], r[1], r[2], r[3], r[4]
    return f"[ ] {pr:>3}. {name}  ({code})  | {mkt} | 시총 {cap}억"

with open(TXT, "w", encoding="utf-8") as f:
    f.write("=" * 70 + "\n")
    f.write("  FILMN9 — DCF 밸류에이션용 애널리스트 리포트 다운로드 체크리스트\n")
    f.write(f"  전체 {len(rows)}개  (코어 {len(core)} / 확장 {len(ext)})\n")
    f.write("  사용법: PDF 받으면 [ ] -> [O] 로 표시\n")
    f.write("=" * 70 + "\n\n")
    f.write(f"■ 코어 (시총 5천억+) — {len(core)}개  ★최우선\n")
    f.write("-" * 70 + "\n")
    for r in core:
        f.write(fmt(r) + "\n")
    f.write("\n")
    f.write(f"■ 확장 (시총 3천억+) — {len(ext)}개\n")
    f.write("-" * 70 + "\n")
    for r in ext:
        f.write(fmt(r) + "\n")

print(f"[OK] txt 저장: {TXT}")
print(f"   코어 {len(core)} / 확장 {len(ext)}")
