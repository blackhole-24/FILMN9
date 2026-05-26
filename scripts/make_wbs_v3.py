"""FILMN9 WBS v3 생성 스크립트"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 색상 상수
C_HEADER_DARK = '1E293B'
C_HEADER_TXT  = 'FFFFFF'
C_DONE        = 'D1FAE5'; C_DONE_TXT   = '065F46'
C_WIP         = 'DBEAFE'; C_WIP_TXT    = '1E40AF'
C_TODO        = 'FFFBEB'; C_TODO_TXT   = '92400E'
C_BACKLOG     = 'F3F4F6'; C_BACKLOG_TXT= '374151'
C_REJECT      = 'F1F5F9'; C_REJECT_TXT = '94A3B8'
C_MUST        = 'FEE2E2'; C_SHOULD     = 'FEF9C3'
C_COULD       = 'E0F2FE'; C_WONT       = 'F3F4F6'

def tborder():
    s = Side(border_style='thin', color='CBD5E1')
    return Border(left=s, right=s, top=s, bottom=s)
def hfill(h): return PatternFill('solid', fgColor=h)
def hfont(c=C_HEADER_TXT, b=True, sz=10): return Font(name='Arial', bold=b, color=c, size=sz)
def bfont(c='1E293B', b=False, sz=9, strike=False): return Font(name='Arial', bold=b, color=c, size=sz, strike=strike)
def cal(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def lal(): return Alignment(horizontal='left', vertical='center', wrap_text=True)

def hdr_row(ws, r, cols, bg=C_HEADER_DARK):
    for ci, v in enumerate(cols, 1):
        c = ws.cell(r, ci, v)
        c.font = hfont(); c.fill = hfill(bg); c.border = tborder(); c.alignment = cal()

def data_row(ws, r, cols, bg, tc='1E293B', b=False, st=False):
    for ci, v in enumerate(cols, 1):
        c = ws.cell(r, ci, v)
        c.font = bfont(tc, b, strike=st); c.fill = hfill(bg); c.border = tborder(); c.alignment = lal()

def sec_row(ws, r, nc, label, bg='F1F5F9', tc='334155'):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(r, 1, label)
    c.font = hfont(tc, True, 9); c.fill = hfill(bg); c.border = tborder(); c.alignment = lal()

STATUS_STYLE = {
    '완료':   (C_DONE,    C_DONE_TXT,    False, False),
    '진행중': (C_WIP,     C_WIP_TXT,     False, False),
    '할 일':  (C_TODO,    C_TODO_TXT,    False, False),
    '백로그': (C_BACKLOG, C_BACKLOG_TXT, False, False),
    '기각':   (C_REJECT,  C_REJECT_TXT,  False, True),
    '보류':   (C_REJECT,  C_REJECT_TXT,  False, False),
}

# ──────────────────────── WBS 시트 ────────────────────────────────────────────
ws1 = wb.active; ws1.title = 'WBS'
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'A3'

ws1.merge_cells('A1:L1')
t = ws1['A1']
t.value = 'FILMN9  WBS v3  —  2026-04-22 ~ 06-20  |  마지막 킥 포함 (Sprint 6~9)'
t.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
t.fill = hfill('0F172A'); t.alignment = cal()
ws1.row_dimensions[1].height = 28

HDR = ['No','스프린트','업무그룹','태스크명','상세 내용','담당','우선순위','상태','공수(일)','시작일','완료일','비고']
hdr_row(ws1, 2, HDR); ws1.row_dimensions[2].height = 20
for i, w in enumerate([5,10,16,28,42,10,8,9,7,11,11,20], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

WBS = [
    ('── 킥오프  04-22 ~ 04-28', 'F0FDF4'),
    (1,'킥오프','킥오프','팀 빌딩 & 헌장 작성','팀명·역할·그라운드룰, 주요기능 논의','전원','Must','완료',1,'2026-04-22','2026-04-22',''),
    (2,'킥오프','킥오프','주제 선정','팀 내부 규칙 설정','전원','Must','완료',2,'2026-04-22','2026-04-23',''),
    (3,'킥오프','킥오프','스프린트 계획서 확정','5스프린트 계획 확정','전원','Must','완료',3,'2026-04-22','2026-04-24',''),
    (4,'킥오프','기획','서비스 범위·페르소나 설정(1차)','팀 공유 노션 정리','손소희','Must','완료',1,'2026-04-27','2026-04-27',''),
    (5,'킥오프','기획','WBS 작성(1차)','','전원','Must','완료',7,'2026-04-22','2026-04-28',''),

    ('── Sprint 1  기획  04-27 ~ 04-28', 'FFF7ED'),
    (6,'Sprint 1','기획','기획서 작성','','전원','Must','완료',2,'2026-04-27','2026-04-28',''),
    (7,'Sprint 1','기획','WBS 정식본','기능 상세·산출물·담당자','전원','Must','완료',2,'2026-04-27','2026-04-28',''),
    (8,'Sprint 1','기획','GitHub 저장소 생성','','이형주','Must','완료',1,'2026-04-27','2026-04-28',''),
    (9,'Sprint 1','기획','요구사항 정의서','개요·배경·시스템·데이터','손소희','Must','완료',2,'2026-04-27','2026-04-28',''),

    ('── Sprint 2  개발환경 / PoC / 1차 멘토링  04-29 ~ 05-09', 'F5F3FF'),
    (10,'Sprint 2','개발환경','Python 가상환경 + 패키지','miniconda3, FastAPI, langchain, pykrx','전원','Must','완료',1,'2026-04-30','2026-04-30',''),
    (11,'Sprint 2','DB 설계','SQLite 메인 DB 스키마','corp_info/financials/ohlcv','전원','Must','완료',9,'2026-04-30','2026-05-08','원본 MySQL → SQLite'),
    (12,'Sprint 2','데이터수집','기업 재무제표 일괄 수집(DART)','BS·PL·CF·주석','전원','Must','완료',9,'2026-04-30','2026-05-08',''),
    (13,'Sprint 2','데이터수집','ECOS API Rf 국고채 수집','3년물','고휘주','Must','완료',3,'2026-04-30','2026-05-08',''),
    (14,'Sprint 2','데이터수집','KRX 시가총액 자동 수집','주가×발행주식수','유태상','Must','완료',3,'2026-04-30','2026-05-08',''),
    (15,'Sprint 2','Valuation PoC','WACC Python 계산 로직','Rf·β·ERP·Kd → WACC','유태상','Must','완료',3,'2026-05-06','2026-05-08',''),
    (16,'Sprint 2','Valuation PoC','DCF 3-Way 시나리오','Bear/Base/Bull','전원','Must','완료',3,'2026-05-06','2026-05-08',''),
    (17,'Sprint 2','Valuation PoC','멀티플 계산(EV/EBITDA·P/E)','','전원','Must','완료',3,'2026-05-06','2026-05-08',''),
    (18,'Sprint 2','멘토링','1차 멘토링 (5/9)','페르소나 피벗: 회계법인→일반투자자','전원','Must','완료',1,'2026-05-09','2026-05-09','변곡점'),

    ('── Sprint 3  피벗 적용 / PoC 데모  05-12 ~ 05-26', 'ECFDF5'),
    (19,'Sprint 3','ERD·아키텍처','ERD v3 FINAL (15테이블)','SQLite 14 + MongoDB 1','이형주','Must','완료',7,'2026-05-12','2026-05-18',''),
    (20,'Sprint 3','DB 구축','MongoDB Atlas + 142개 히스토리','history 컬렉션','유태상','Must','완료',5,'2026-05-15','2026-05-19',''),
    (21,'Sprint 3','DB 구축','SQLite 22개 기업 + OHLCV 56,069행','pykrx 일봉','유태상','Must','완료',7,'2026-05-13','2026-05-19',''),
    (22,'Sprint 3','백엔드 API','FastAPI 6엔드포인트','company·financials·overview·ohlcv·search','유태상','Must','완료',5,'2026-05-17','2026-05-21',''),
    (23,'Sprint 3','백엔드 API','/api/valuation/{code} 신규','6파일 구조 파싱, _convert_multi()','이형주','Must','완료',2,'2026-05-22','2026-05-22',''),
    (24,'Sprint 3','데이터','RAG 임베딩 (2,596청크 18GB)','text-embedding-3-small, ChromaDB','유태상','Must','완료',9,'2026-05-12','2026-05-20',''),
    (25,'Sprint 3','프론트엔드','v0.dev React 5탭 생성','Streamlit 대체','손소희','Must','완료',1,'2026-05-22','2026-05-22',''),
    (26,'Sprint 3','프론트엔드','Tab1 히스토리 브리핑 실데이터','JSON 폴백, 3종목 카드 렌더링','이형주','Must','완료',1,'2026-05-22','2026-05-22',''),
    (27,'Sprint 3','프론트엔드','Tab2 밸류에이션 실데이터 자동전환','vd fallback, ✓실데이터 배지','이형주','Must','완료',1,'2026-05-22','2026-05-23',''),
    (28,'Sprint 3','프론트엔드','삼성전기 DCF Invalid 경고 배너','노란 경고박스, dcf_valid 플래그','이형주','Must','완료',0,'2026-05-23','2026-05-23',''),
    (29,'Sprint 3','Sankey','Sankey 3종목 단위 오류 수정','원→백만원, NAVER 영업비용 처리','이형주','Must','완료',0,'2026-05-23','2026-05-23',''),
    (30,'Sprint 3','Sankey','Sankey 디자인 최종 수정','디자인 확정 후 재생성','이형주','Should','할 일',1,'2026-05-23','2026-05-24','저녁 이후'),
    (31,'Sprint 3','데모 준비','AI 기능 4종 설계서','종합분석·Q&A·비교·모닝루틴','이형주','Must','할 일',1,'2026-05-23','2026-05-23',''),
    (32,'Sprint 3','데모 준비','전체 플로우 점검 (3종목 × 전탭)','버그 발견 즉시 수정','전원','Must','할 일',1,'2026-05-23','2026-05-24',''),
    (33,'Sprint 3','데모 준비','시연 스크립트 최종 확인','5분 스크립트 + 예상질문 30개','이형주','Must','할 일',1,'2026-05-24','2026-05-24',''),
    (34,'Sprint 3','데모 준비','삼성전기 DCF 괴리 설명 멘트','-93% 업사이드 사전 설명','이형주','Must','할 일',1,'2026-05-24','2026-05-24',''),
    (35,'Sprint 3','데모 준비','최종 리허설','오전 9-11시 환경 점검','전원','Must','할 일',1,'2026-05-25','2026-05-25',''),
    (36,'Sprint 3','데모','PoC 데모 — KPMG AI Lab','5/26 데모','전원','Must','할 일',1,'2026-05-26','2026-05-26',''),

    ('── Sprint 4  2차 배포 / AX 강화  05-26 ~ 06-05', 'FFF1F2'),
    (37,'Sprint 4','히스토리 브리핑','브리핑 생성 전체 파이프라인','DART+네이버뉴스+yfinance','이형주,고휘주','Must','할 일',11,'2026-05-26','2026-06-05',''),
    (38,'Sprint 4','기업개요 차별성','① 감사의견 자동 추출','적정/한정/거절','이형주,고휘주','Must','할 일',3,'2026-05-26','2026-06-05',''),
    (39,'Sprint 4','기업개요 차별성','② 리스크 자동 감지','소송·신용등급·CB발행','이형주,고휘주','Must','할 일',3,'2026-05-26','2026-06-05',''),
    (40,'Sprint 4','Valuation','4-Way 내재가치 비교 바차트','DCF·PER·PBR·EV/EBITDA','유태상,손소희','Must','할 일',3,'2026-05-26','2026-06-05',''),
    (41,'Sprint 4','Valuation','AI 고평가/저평가 진단 LLM','gpt-5-mini 종합 근거','전원','Must','할 일',3,'2026-05-26','2026-06-05',''),
    (42,'Sprint 4','AX 강화','A1 "왜 이 가격?" AI 자가해설','ChromaDB 재활용','고휘주','Must','할 일',3,'2026-05-28','2026-05-30','KPMG 평가핵심'),
    (43,'Sprint 4','AX 강화','A2 오늘의 한 줄 브리핑','일배치 + 알림, lock-in','유태상','Must','할 일',6,'2026-05-29','2026-06-03',''),
    (44,'Sprint 4','AX 강화','C1 LangGraph 멀티 에이전트','KPMG AX 평가 핵심','전팀','Must','할 일',7,'2026-05-30','2026-06-05',''),
    (45,'Sprint 4','배포','AWS EC2 (서울) 2차 배포','발표·테스트 시만 가동','전원','Must','할 일',3,'2026-06-03','2026-06-05','예산 절감'),
    (46,'Sprint 4','멘토링','3차 멘토링 (5/30)','','전원','Must','할 일',1,'2026-05-30','2026-05-30',''),

    ('── Sprint 5  QA / 최종 배포 / 발표 준비  06-08 ~ 06-19', 'F0FDF4'),
    (47,'Sprint 5','QA','테스트 케이스 작성 (30개+)','기업개요·히스토리·WACC·DCF','전원','Must','할 일',4,'2026-06-08','2026-06-11',''),
    (48,'Sprint 5','QA','통합 테스트 + 버그 수정','','전원','Must','할 일',6,'2026-06-10','2026-06-15',''),
    (49,'Sprint 5','AX 강화','A3 AI 종목 비교 챗봇','Agent 패턴','유태상,고휘주','Could','할 일',5,'2026-06-08','2026-06-12',''),
    (50,'Sprint 5','AX 강화','A4 호재·악재 자동 신호등','페르소나 직격','유태상','Could','할 일',3,'2026-06-15','2026-06-17',''),
    (51,'Sprint 5','배포·산출물','최종 서비스 안정화 + 3차 배포','','전원','Must','할 일',2,'2026-06-15','2026-06-16',''),
    (52,'Sprint 5','배포·산출물','전체 산출물 패키징·제출','기획서·WBS·코드·URL','전원','Must','할 일',2,'2026-06-17','2026-06-18',''),
    (53,'Sprint 5','발표 준비','최종 발표 PPT','서비스 데모·Valuation·차별점','전원','Must','할 일',8,'2026-06-10','2026-06-17',''),
    (54,'Sprint 5','발표 준비','발표 리허설 + 레슨런','','전원','Must','할 일',1,'2026-06-19','2026-06-19',''),
    (55,'Sprint 5','멘토링','4차 멘토링 (6/13)','','전원','Must','할 일',1,'2026-06-13','2026-06-13',''),

    ('── 최종 발표  06-20', 'F0FDF4'),
    (56,'최종발표','발표','최종 발표 및 수료식','','전원','Must','할 일',1,'2026-06-20','2026-06-20',''),

    ('── Sprint 6  마지막 킥 ①  메인화면 재설계  (데모 이후 추진)', 'FFF7ED'),
    (57,'Sprint 6','메인화면','UX 여정 재설계 기획','페르소나 행동 패턴, 사용자 여정 재정의','손소희','Must','백로그',3,'','','마지막 킥'),
    (58,'Sprint 6','메인화면','Tab0 시황 대시보드 구현','코스피/코스닥 지수, 업종별 등락 히트맵','이형주,유태상','Must','백로그',5,'','','마지막 킥'),
    (59,'Sprint 6','메인화면','오늘의 시황 요약 카드','상승/하락 상위 종목, 시가총액 TOP10','이형주','Should','백로그',3,'','','마지막 킥'),
    (60,'Sprint 6','메인화면','"요즘 뜨는 산업" 배너','뉴스 기반 AI 키워드 추출','고휘주','Could','백로그',3,'','','마지막 킥'),
    (61,'Sprint 6','메인화면','계절성 알림 모듈','이맘때 오르는 섹터 패턴','유태상','Could','백로그',3,'','','마지막 킥'),

    ('── Sprint 7  마지막 킥 ②  산업 분류 DB  (데모 이후 추진)', 'EFF6FF'),
    (62,'Sprint 7','산업분류','WICS 기준 2,695개 기업 산업 태깅','KRX 공개 데이터 + DART 업종코드','유태상','Must','백로그',10,'','','마지막 킥 핵심 인프라'),
    (63,'Sprint 7','산업분류','유사어 사전(synonym dict) 구축','방산=국방=항공=방위산업 매핑','손소희','Must','백로그',5,'','','마지막 킥'),
    (64,'Sprint 7','산업분류','대장주 선정 로직 (시가총액 기준)','산업별 TOP N 자동 선정','유태상','Must','백로그',3,'','','마지막 킥'),
    (65,'Sprint 7','산업분류','산업별 재무우량/미래전망 분류','재무지표 기반 그룹 알고리즘','유태상,손소희','Could','백로그',7,'','','마지막 킥'),

    ('── Sprint 8  마지막 킥 ③  업종 검색 기능  (데모 이후 추진)', 'FDF4FF'),
    (66,'Sprint 8','업종검색','업종 키워드 검색 API','유사어 매핑, LIKE 검색','유태상','Must','백로그',5,'','','마지막 킥'),
    (67,'Sprint 8','업종검색','업종 결과 화면 (대장주+리스트)','대장주 TOP3 카드 + 전체 스크롤','이형주,손소희','Must','백로그',5,'','','마지막 킥'),
    (68,'Sprint 8','업종검색','그룹 탭 (대장주/재무우량/전망)','','손소희','Should','백로그',3,'','','마지막 킥'),
    (69,'Sprint 8','업종검색','통합 검색창 (종목+업종)','검색창 하나로 종목코드·회사명·업종','이형주','Must','백로그',3,'','','마지막 킥'),

    ('── Sprint 9  마지막 킥 ④  AI 시황 브리핑  (데모 이후 추진)', 'FAFAFA'),
    (70,'Sprint 9','AI 시황','뉴스 수집 일배치 파이프라인','RSS / 네이버경제 / 한경 / 매경','유태상','Must','백로그',5,'','','마지막 킥'),
    (71,'Sprint 9','AI 시황','경제 지표 해석 엔진','금리↑=성장주 부담, 유가↑=정유 수혜 등','이형주,유태상','Must','백로그',7,'','','마지막 킥'),
    (72,'Sprint 9','AI 시황','"이번 주 시장 요약" 카드','주 1회 AI 생성, 메인화면 노출','이형주','Should','백로그',3,'','','마지막 킥'),
    (73,'Sprint 9','AI 시황','계절성·패턴 분석 (과거 데이터)','분기·월별 섹터 강세 패턴','유태상','Could','백로그',10,'','','마지막 킥'),
]

r = 3
for item in WBS:
    if isinstance(item[0], str):
        sec_row(ws1, r, 12, item[0], item[1])
        ws1.row_dimensions[r].height = 16
    else:
        no,sp,grp,task,desc,asgn,pri,sts,eff,st,en,note = item
        bg,tc,bd,sk = STATUS_STYLE.get(sts,(C_TODO,C_TODO_TXT,False,False))
        data_row(ws1, r, [no,sp,grp,task,desc,asgn,pri,sts,eff if eff else '',st,en,note], bg, tc, bd, sk)
        ws1.row_dimensions[r].height = 30
    r += 1

# 요약 공식 행
ws1.merge_cells(f'A{r}:L{r}')
c = ws1.cell(r, 1, f'=COUNTA(H3:H{r-1})&" 개 태스크 (완료: "&COUNTIF(H3:H{r-1},"완료")&" / 진행중: "&COUNTIF(H3:H{r-1},"진행중")&" / 할 일: "&COUNTIF(H3:H{r-1},"할 일")&" / 백로그: "&COUNTIF(H3:H{r-1},"백로그")&")"')
c.font = Font(name='Arial', bold=True, size=9, color='1E293B')
c.fill = hfill('F8FAFC'); c.border = tborder(); c.alignment = lal()

# ──────────────────────── 백로그 시트 ─────────────────────────────────────────
ws2 = wb.create_sheet('백로그')
ws2.sheet_view.showGridLines = False; ws2.freeze_panes = 'A3'

ws2.merge_cells('A1:K1')
t2 = ws2['A1']
t2.value = 'FILMN9  Product Backlog  —  PoC 완료 + Sprint 4~5 + 마지막 킥 (Sprint 6~9)'
t2.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
t2.fill = hfill('0F172A'); t2.alignment = cal()
ws2.row_dimensions[1].height = 28

BL_HDR = ['ID','기능명','사용자 스토리','카테고리','MoSCoW','스프린트','공수(일)','상태','검증 기준','담당','비고']
hdr_row(ws2, 2, BL_HDR); ws2.row_dimensions[2].height = 20
for i, w in enumerate([7,22,46,12,8,10,7,9,28,12,14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

BACKLOG = [
    ('── 완료된 핵심 기능 (PoC 데모 기준)', 'D1FAE5'),
    ('BL-001','주가 차트 (캔들+거래량+MA)','투자자로서 주가 흐름을 보고 싶다','기업개요','Must','Sprint 3',8,'완료','캔들+MA5/20/60+볼린저 렌더링','유태상',''),
    ('BL-002','기업 기본 정보 표시','투자자로서 기업 개요·경영인·주주를 보고 싶다','기업개요','Must','Sprint 3',5,'완료','3종목 정상 표시','이형주,고휘주',''),
    ('BL-003','AI 히스토리 브리핑 (Tab1)','투자자로서 기업 역사·BM을 AI 요약으로 보고 싶다','기업개요','Must','Sprint 3',3,'완료','3종목 카드 렌더링','이형주',''),
    ('BL-004','Sankey 손익 흐름도','투자자로서 돈의 흐름을 시각적으로 보고 싶다','기업개요','Should','Sprint 3',4,'완료','3종목 단위 정상 표시 (조/억)','손소희','단위 오류 수정 완료'),
    ('BL-005','DCF 밸류에이션 (Tab2)','투자자로서 DCF 적정가를 자동 계산받고 싶다','밸류에이션','Must','Sprint 3',11,'완료','Bear/Base/Bull 3시나리오','유태상,손소희',''),
    ('BL-006','WACC 계산 자동화','투자자로서 WACC를 자동 계산받고 싶다','밸류에이션','Must','Sprint 3',5,'완료','수치 정상 표시','유태상',''),
    ('BL-007','민감도 히트맵 (WACC×g)','투자자로서 민감도 분석을 보고 싶다','밸류에이션','Must','Sprint 3',3,'완료','3×3 히트맵 Base셀 강조','유태상,손소희',''),
    ('BL-008','DCF Invalid 경고 배너','투자자로서 DCF 신뢰불가 시 경고를 받고 싶다','밸류에이션','Must','Sprint 3',1,'완료','삼성전기 경고 배너 정상','이형주',''),

    ('── Sprint 4 (05-26 ~ 06-05)', 'DBEAFE'),
    ('BL-009','브리핑 파이프라인 완성','투자자로서 어떤 종목이든 AI 브리핑을 받고 싶다','기업개요','Must','Sprint 4',11,'할 일','22개 → 전체 종목 생성 가능','이형주,고휘주',''),
    ('BL-010','감사의견 자동 추출','투자자로서 감사의견(적정/한정)을 즉시 알고 싶다','차별성','Must','Sprint 4',3,'할 일','DART 파싱 정확도 90%+','이형주,고휘주',''),
    ('BL-011','리스크 자동 감지','투자자로서 소송·CB 발행 위험신호를 알고 싶다','차별성','Must','Sprint 4',3,'할 일','공시 파싱 후 플래그 표시','이형주,고휘주',''),
    ('BL-012','4-Way 내재가치 비교 바차트','투자자로서 4개 방법론 적정가를 비교하고 싶다','밸류에이션','Must','Sprint 4',3,'할 일','4개 바 정상 렌더링','유태상,손소희',''),
    ('BL-013','AI 고평가/저평가 진단','투자자로서 이 주식이 싸고 비싼지 AI 해석을 원한다','밸류에이션','Must','Sprint 4',3,'할 일','LLM 진단 문장 출력','전원',''),
    ('BL-014','A2 오늘의 한 줄 브리핑','투자자로서 매일 관심 종목 한 줄 요약을 받고 싶다','AX 강화','Must','Sprint 4',6,'할 일','일배치 + UI 노출','유태상','lock-in 핵심'),
    ('BL-015','A1 "왜 이 가격?" 자가해설','투자자로서 DCF 결과 AI 설명을 듣고 싶다','AX 강화','Must','Sprint 4',3,'할 일','챗봇 응답 정상','고휘주','KPMG 평가핵심'),
    ('BL-016','C1 LangGraph 멀티 에이전트','투자자로서 복합 질문에 단계 추론 답변을 원한다','AX 강화','Must','Sprint 4',7,'할 일','3단계 이상 추론 시연','전팀','KPMG AX 핵심'),
    ('BL-017','AWS EC2 2차 배포','팀으로서 외부 접속 가능한 URL이 필요하다','인프라','Must','Sprint 4',3,'할 일','URL 접속 정상','전원',''),

    ('── Sprint 5 (06-08 ~ 06-19)', 'D1FAE5'),
    ('BL-018','통합 테스트 30개+ 케이스','팀으로서 서비스 품질을 체계적으로 검증하고 싶다','QA','Must','Sprint 5',10,'할 일','Pass율 95%+','전원',''),
    ('BL-019','A3 AI 종목 비교 챗봇','투자자로서 두 종목을 동시에 비교 질문하고 싶다','AX 강화','Could','Sprint 5',5,'할 일','2종목 비교 응답 정상','유태상,고휘주',''),
    ('BL-020','A4 호재·악재 자동 신호등','투자자로서 기업 뉴스의 호악재 여부를 즉시 알고 싶다','AX 강화','Could','Sprint 5',3,'할 일','신호등 UI 표시','유태상',''),
    ('BL-021','최종 산출물 패키징','팀으로서 제출 산출물을 완벽히 정리하고 싶다','산출물','Must','Sprint 5',5,'할 일','제출 체크리스트 100%','전원',''),

    ('── 마지막 킥 — Sprint 6: 메인화면 재설계 (데모 이후)', 'FFF7ED'),
    ('BL-022','UX 여정 재설계','주식 초보로서 처음 와도 무엇을 할지 알 수 있길 원한다','UX','Must','Sprint 6',3,'백로그','사용자 테스트 3명 통과','손소희','마지막 킥'),
    ('BL-023','Tab0 시황 대시보드','일반 투자자로서 오늘 시장 상황을 첫 화면에서 보고 싶다','메인화면','Must','Sprint 6',5,'백로그','코스피/코스닥 + 업종 히트맵','이형주,유태상','마지막 킥'),
    ('BL-024','오늘의 시황 요약 카드','투자자로서 상승/하락 상위 종목을 한눈에 보고 싶다','메인화면','Should','Sprint 6',3,'백로그','TOP10 정상 표시','이형주','마지막 킥'),
    ('BL-025','"요즘 뜨는 산업" 배너','투자자로서 지금 주목받는 산업을 알고 싶다','메인화면','Could','Sprint 6',3,'백로그','AI 키워드 배너 표시','고휘주','마지막 킥'),
    ('BL-026','계절성 알림 모듈','투자자로서 이맘때 오르는 섹터 패턴을 알고 싶다','메인화면','Could','Sprint 6',3,'백로그','과거 패턴 기반 알림','유태상','마지막 킥'),

    ('── 마지막 킥 — Sprint 7: 산업 분류 DB (데모 이후)', 'EFF6FF'),
    ('BL-027','2,695개 기업 WICS 태깅','플랫폼으로서 모든 상장사에 산업 태그를 부여하고 싶다','인프라 DB','Must','Sprint 7',10,'백로그','태깅 완료율 95%+','유태상','마지막 킥 핵심'),
    ('BL-028','유사어 사전 구축','투자자로서 비슷한 말로도 검색 가능하길 원한다','검색','Must','Sprint 7',5,'백로그','50개+ 산업 유사어 커버','손소희','마지막 킥'),
    ('BL-029','대장주 자동 선정 로직','투자자로서 산업별 대표 종목을 자동으로 보고 싶다','검색','Must','Sprint 7',3,'백로그','시가총액 기준 TOP3','유태상','마지막 킥'),
    ('BL-030','산업별 재무우량/전망 분류','투자자로서 같은 산업에서도 그룹별로 보고 싶다','검색','Could','Sprint 7',7,'백로그','그룹 분류 알고리즘 동작','유태상,손소희','마지막 킥'),

    ('── 마지막 킥 — Sprint 8: 업종 검색 (데모 이후)', 'FDF4FF'),
    ('BL-031','업종 키워드 검색 API','"반도체" 입력으로 관련 종목 전체를 보고 싶다','검색','Must','Sprint 8',5,'백로그','검색결과 정상 반환','유태상','마지막 킥'),
    ('BL-032','업종 결과 화면','검색 후 대장주와 전체 리스트를 보고 싶다','검색 UI','Must','Sprint 8',5,'백로그','UI 렌더링 정상','이형주,손소희','마지막 킥'),
    ('BL-033','통합 검색창 (종목+업종)','검색창 하나로 모든 것을 찾고 싶다','검색 UI','Must','Sprint 8',3,'백로그','3가지 타입 구분 표시','이형주','마지막 킥'),

    ('── 마지막 킥 — Sprint 9: AI 시황 브리핑 (데모 이후)', 'FAFAFA'),
    ('BL-034','뉴스 수집 일배치 파이프라인','플랫폼으로서 매일 경제 뉴스를 자동 수집하고 싶다','AI 시황','Must','Sprint 9',5,'백로그','매일 100건+ 수집','유태상','마지막 킥'),
    ('BL-035','경제 지표 해석 엔진','주식 초보로서 뉴스가 주가에 미치는 영향을 알고 싶다','AI 시황','Must','Sprint 9',7,'백로그','5개+ 지표 섹터 연동','이형주,유태상','마지막 킥'),
    ('BL-036','주간 시장 요약 카드','바쁜 투자자로서 이번 주 시장을 1분에 파악하고 싶다','AI 시황','Should','Sprint 9',3,'백로그','주 1회 AI 생성 정상','이형주','마지막 킥'),
    ('BL-037','계절성·패턴 분석','투자자로서 이맘때 오르는 주식을 추천받고 싶다','AI 시황','Could','Sprint 9',10,'백로그','과거 3년 패턴 기반','유태상','마지막 킥'),
]

r2 = 3
for item in BACKLOG:
    if isinstance(item[0], str) and len(item) == 2:
        sec_row(ws2, r2, 11, item[0], item[1])
        ws2.row_dimensions[r2].height = 16
    else:
        bid,name,story,cat,moscow,sp,eff,sts,crit,asgn,note = item
        bg  = {'완료':C_DONE,'할 일':C_TODO,'백로그':C_BACKLOG}.get(sts,C_BACKLOG)
        tc  = {'완료':C_DONE_TXT,'할 일':C_TODO_TXT,'백로그':C_BACKLOG_TXT}.get(sts,C_BACKLOG_TXT)
        data_row(ws2, r2, [bid,name,story,cat,moscow,sp,eff,sts,crit,asgn,note], bg, tc)
        ws2.row_dimensions[r2].height = 36
    r2 += 1

# ──────────────────────── 스프린트 계획 시트 ──────────────────────────────────
ws3 = wb.create_sheet('스프린트 계획')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:H1')
t3 = ws3['A1']
t3.value = 'FILMN9  Sprint Overview  —  PoC ~ 마지막 킥 전체 로드맵'
t3.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
t3.fill = hfill('0F172A'); t3.alignment = cal()
ws3.row_dimensions[1].height = 28

SP_HDR = ['스프린트','기간','목표','주요 태스크','완료 기준 (DoD)','상태','비고','마지막 킥']
hdr_row(ws3, 2, SP_HDR); ws3.row_dimensions[2].height = 20
for i, w in enumerate([13,22,30,48,32,9,18,8], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

SPRINTS = [
    ('킥오프','2026-04-22 ~ 04-28','팀 빌딩, 주제 선정, WBS 1차','헌장서·WBS·페르소나·저장소 생성','헌장 서명, GitHub 개설, WBS 등록','완료','',''),
    ('Sprint 1\n기획','2026-04-27 ~ 04-28','기획서 & 요구사항 완성','기획서·요구사항정의서·WBS 정식본','기획서 드라이브 등록, 팀 전원 리뷰','완료','',''),
    ('Sprint 2\n개발환경/PoC','2026-04-29 ~ 05-09','개발환경 + Valuation PoC 완성','가상환경·DB·DART·WACC·DCF·멀티플','3~5개 기업 Valuation 계산 정합성 확인','완료','1차 멘토링(5/9) 피벗',''),
    ('Sprint 3\nPoC 데모','2026-05-12 ~ 05-26','ERD·프론트·3종목 실데이터·5/26 데모','ERD v3·API·React(v0)·Tab1/Tab2·Sankey','3종목 × 전탭 정상 렌더링, 리허설 완료','진행중','5/26 데모 예정',''),
    ('Sprint 4\n2차 배포/AX','2026-05-26 ~ 06-05','브리핑 파이프라인 + AX 3종 + AWS','A1자가해설·A2한줄브리핑·C1LangGraph·AWS','AX 기능 시연 가능, AWS URL 접속 정상','할 일','3차 멘토링(5/30)',''),
    ('Sprint 5\nQA/최종배포','2026-06-08 ~ 06-19','QA 완료, 3차 배포, 발표 준비','TC 30개+·통합테스트·발표PPT·산출물 패키징','Pass율 95%+, 제출 100%, 리허설 1회','할 일','4차 멘토링(6/13)',''),
    ('최종 발표','2026-06-20','KPMG AI Lab 최종 발표 및 수료','서비스 데모·발표 Q&A','발표 완료, 수료','할 일','',''),
    ('Sprint 6\n메인화면','데모 이후 추진','UX 재설계 + Tab0 시황 대시보드','UX 여정·코스피/코스닥·업종 히트맵·시황카드·뜨는산업 배너','주식 초보 사용자 테스트 3명 통과','백로그','마지막 킥 ①','O'),
    ('Sprint 7\n산업분류','데모 이후 추진','2,695개 기업 산업 분류 DB 구축','WICS 태깅·유사어 사전·대장주 선정·재무그룹 분류','전체 상장사 태깅 완료율 95%+','백로그','마지막 킥 ② — 핵심 인프라','O'),
    ('Sprint 8\n업종검색','데모 이후 추진','업종 키워드 검색 기능 신설','검색 API·대장주+리스트화면·그룹탭·통합검색창','"방산" → 한화에어로스페이스 등 정상','백로그','마지막 킥 ③','O'),
    ('Sprint 9\nAI 시황','데모 이후 추진','AI 뉴스 시황 브리핑 + 계절성 분석','뉴스 일배치·해석엔진·주간요약·패턴분석','매일 뉴스 수집 + AI 해석 정상 출력','백로그','마지막 킥 ④','O'),
]

STATUS_BG3  = {'완료':C_DONE,'진행중':C_WIP,'할 일':C_TODO,'백로그':C_BACKLOG}
STATUS_TXT3 = {'완료':C_DONE_TXT,'진행중':C_WIP_TXT,'할 일':C_TODO_TXT,'백로그':C_BACKLOG_TXT}

for ri, sp in enumerate(SPRINTS, 3):
    bg  = STATUS_BG3.get(sp[5], C_BACKLOG)
    tc  = STATUS_TXT3.get(sp[5], C_BACKLOG_TXT)
    for ci, v in enumerate(sp, 1):
        c = ws3.cell(ri, ci, v)
        c.font = bfont(tc, b=(ci==1)); c.fill = hfill(bg)
        c.border = tborder(); c.alignment = lal()
    ws3.row_dimensions[ri].height = 56

# ──────────────────────── 저장
out = 'C:/Users/Admin/Downloads/FILMN9_WBS_v3.xlsx'
wb.save(out)
print('완료:', out)
