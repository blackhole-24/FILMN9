# -*- coding: utf-8 -*-
"""FILMN9 WBS v5 — 스프린트 4·5만 (2차·3차 배포)"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "FILMN9 WBS v5"

# 색상
HEADER_BG = "1F4E3D"
SPRINT_BG = "B8E0C9"
GOAL_BG   = "FFF3E0"   # 목표 박스 — 살구
A_HJ  = "2C3E50"; A_HW  = "C0392B"; A_YT  = "27AE60"; A_ALL = "8E44AD"
S_OK   = "C8E6C9"; S_DOING = "FFF59D"; S_TODO = "BBDEFB"
S_CHK  = "FFCC80"; S_REJ  = "FFCDD2"; S_HOLD = "E1BEE7"; S_KEY = "F8BBD0"

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fill(c): return PatternFill("solid", fgColor=c)
def ac(n):
    if not n: return None
    if "전원" in n: return A_ALL
    if "이형주" in n: return A_HJ
    if "고휘주" in n: return A_HW
    if "유태상" in n: return A_YT
    return None
def sc(s):
    if not s: return None
    if "완료" in s: return S_OK
    if "진행" in s: return S_DOING
    if "할 일" in s or "할일" in s: return S_TODO
    if "확인" in s: return S_CHK
    if "기각" in s: return S_REJ
    if "보류" in s: return S_HOLD
    if "핵심" in s: return S_KEY
    return None

# ── 상단 헤더 ───────────────────────────────────────
ws["A1"] = "FILMN9 WBS v5 — 스프린트 4 & 5 (2차·3차 배포 + 최종 발표)"
ws["A1"].font = Font(name="맑은 고딕", size=15, bold=True, color="FFFFFF")
ws["A1"].fill = fill(HEADER_BG)
ws.merge_cells("A1:I1")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

# ── 목표 박스 ──────────────────────────────────────
goals = [
    ("🎯 스프린트 4 (05/26~06/05) 목표 — 2차 배포",
     "1) 1차 배포 데모 화면 토대로 보완 + 2,695개 기업 전체 자동화·정확성 확보\n"
     "2) AWS 환경 이전 (S3·EC2·DB) — 클라우드 정식 배포\n"
     "3) v0.dev 기반 UI/UX 고도화 — 클릭 동선·로딩 속도·반응성 개선\n"
     "4) [각 파트] 히스토리브리핑(휘주) · 주가차트/재무(형주) · 밸류에이션(태상)\n"
     "5) Plan B: 시간 부족 시 5차 스프린트 마지막 발표 전까지 연장"),
    ("🎯 스프린트 5 (06/08~06/19) 목표 — 3차 배포 + 최종 완성",
     "1) 메인 화면 신규 구현 (산업 분류 + 모닝루틴 위젯)\n"
     "2) Tab3 AI 종합분석 + RAG 챗봇 고도화 (정식 Chroma 엔진)\n"
     "3) 2,695개 전 종목 정확성 최종 검증\n"
     "4) AWS 배포 마무리 + 도메인·SSL\n"
     "5) 최종 발표(06/20) 자료·리허설"),
]
row = 2
for title, body in goals:
    ws.cell(row=row, column=1, value=title)
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=9)
    cell = ws.cell(row=row, column=1)
    cell.font = Font(name="맑은 고딕", size=11, bold=True, color="1F4E3D")
    cell.fill = fill(SPRINT_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1
    ws.cell(row=row, column=1, value=body)
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=9)
    cell = ws.cell(row=row, column=1)
    cell.font = Font(name="맑은 고딕", size=10, color="1F4E3D")
    cell.fill = fill(GOAL_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 75
    row += 1

# 한 줄 띄우기
row += 1

# ── 컬럼 헤더 ──────────────────────────────────────
headers = ["Group", "Task", "Subtask", "Link", "Assignee", "Start Date", "End Date", "Period", "Status"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    c.fill = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
ws.row_dimensions[row].height = 30
header_row = row
row += 1

# 컬럼 폭
widths = [13, 22, 62, 12, 14, 11, 11, 8, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── 데이터 ─────────────────────────────────────────
# (Group, Task, Subtask, Link, Assignee, Start, End, Period, Status, is_header)
data = [
    # ============ 스프린트 4 ============
    ("스프린트 4", "", "★ 2차 배포 — 2,695개 기업 자동화·정확성 + AWS 이전 + UI 고도화", "", "", "05/26", "06/05", 11, "진행 중", True),

    # ─── 1. 데이터 정확성·자동화 (전 종목 보완) ───
    ("05/26\n~\n06/05",
     "[기업개요·재무]\n전 종목 정확성", "사업보고서 XML 확장 다운로드 1,295 → 2,844개", "[github]", "이형주", "05/26", "05/28", 3, "완료", False),
    ("", "", "XML inline XBRL 파싱 + DB 적재 (1,861개 · 625,218건)", "[github]", "이형주", "05/26", "05/28", 3, "완료", False),
    ("", "", "백엔드 API: 연결/별도 토글 + 음수처리 + 우선정렬", "[github]", "이형주", "05/27", "05/28", 2, "완료", False),
    ("", "", "외감법인 529개 — JSONL+임베딩 활용 보충 (5/27 메모 참조)", "[드라이브]", "이형주\n고휘주", "06/01", "06/05", 5, "할 일", False),
    ("", "", "★ 2,695개 전 종목 재무상태표·손익계산서 100% 자동화", "[github]", "이형주", "06/01", "06/05", 5, "핵심", False),
    ("", "", "OHLCV 자동 갱신 스케줄러 (매일 장 마감 후 자동 fetch)", "[github]", "이형주", "06/02", "06/05", 4, "할 일", False),
    ("", "", "주가 데이터 검증 자동화 (KRX vs DB 일치 확인)", "[드라이브]", "이형주", "06/02", "06/05", 4, "할 일", False),

    # ─── 2. 히스토리 브리핑 전체 확장 ───
    ("", "[기업개요]\nAI 히스토리 브리핑", "briefs_final 2,526개 → 전체 2,695개 종목 확장", "[mongodb]", "고휘주", "05/26", "05/30", 5, "완료", False),
    ("", "", "차별성 ①: 감사의견 자동 추출 (DART 감사보고서)", "[드라이브]", "고휘주", "05/26", "06/05", 11, "할 일", False),
    ("", "", "차별성 ②: 리스크 요소 자동 감지 (소송·신용등급 변동·CB 발행)", "[드라이브]", "고휘주", "05/26", "06/05", 11, "할 일", False),
    ("", "", "차별성 ③: 단일판매 공급계약 공시 자동 추출 (주요사항보고서)", "[드라이브]", "고휘주", "05/26", "06/05", 11, "할 일", False),
    ("", "", "차별성 ④: 주가 영향 요인 AI 분석 (계절·원자재·환율·정책)", "[드라이브]", "고휘주", "05/26", "06/05", 11, "할 일", False),
    ("", "", "★ 2,695개 전 종목 브리핑 신뢰도 등급 자동 분류", "[mongodb]", "고휘주", "06/01", "06/05", 5, "핵심", False),

    # ─── 3. 밸류에이션 1차 통합 ───
    ("", "Valuation\n1차 산출물 통합", "태상님 ui_payload JSON (3종목) 백엔드 통합", "[github]", "유태상\n이형주", "05/26", "05/30", 5, "할 일", False),
    ("", "", "Bear/Base/Bull 카드 실데이터 연결 + 비정상 수치 재검증", "[드라이브]", "유태상\n이형주", "05/26", "05/30", 5, "할 일", False),
    ("", "", "DCF·WACC·멀티플·민감도·토네이도·피어비교 6탭 실데이터 통합", "[드라이브]", "유태상\n이형주", "05/26", "06/02", 8, "할 일", False),
    ("", "", "가치진단 5종 추가 (EPV·성장 프리미엄·시장함의 성장률·기대격차·안전마진)", "[드라이브]", "유태상", "05/26", "06/05", 11, "할 일", False),
    ("", "", "★ 2,695개 전 종목 적정주가 자동 산출 파이프라인 구축", "[github]", "유태상", "06/01", "06/05", 5, "핵심", False),

    # ─── 4. AWS 이전 ★★★ ───
    ("", "AWS 클라우드 배포\n(★ 2차 배포 핵심)", "AWS 계정 셋업 + IAM·VPC 기본 구성", "[aws]", "이형주", "05/28", "05/30", 3, "할 일", False),
    ("", "", "RDS (또는 EC2 + SQLite) 데이터베이스 이전 — 1,861개 종목", "[aws]", "이형주", "05/30", "06/02", 4, "할 일", False),
    ("", "", "S3 버킷 셋업 — XML·JSONL·임베딩 백업·서빙", "[aws]", "이형주", "05/30", "06/02", 4, "할 일", False),
    ("", "", "EC2 인스턴스 — FastAPI 백엔드 배포 (uvicorn + nginx)", "[aws]", "이형주", "06/01", "06/04", 4, "할 일", False),
    ("", "", "EC2 또는 Vercel — Next.js 프런트 배포", "[aws]", "이형주", "06/01", "06/04", 4, "할 일", False),
    ("", "", "MongoDB Atlas 화이트리스트에 AWS IP 추가", "[mongodb]", "이형주", "06/02", "06/03", 2, "할 일", False),
    ("", "", "환경변수·SECRET (DART/OpenAI/Mongo) AWS Secrets Manager", "[aws]", "이형주", "06/02", "06/03", 2, "할 일", False),
    ("", "", "도메인·SSL 인증서 (Route53 + ACM) 적용 (선택)", "[aws]", "이형주", "06/04", "06/05", 2, "할 일", False),

    # ─── 5. UI/UX 고도화 v0.dev ───
    ("", "UI/UX 고도화\n(v0.dev 활용)", "v0.dev 메인화면 신규 디자인 프롬프트 작성·생성", "[v0.dev]", "이형주", "05/28", "06/02", 6, "할 일", False),
    ("", "", "클릭 동선 분석 + 페이지 진입~상세 단계 최적화", "[드라이브]", "이형주\n전원", "05/28", "06/02", 6, "할 일", False),
    ("", "", "로딩 속도 최적화 (스켈레톤 UI + lazy load + 캐싱)", "[github]", "이형주", "05/30", "06/04", 6, "할 일", False),
    ("", "", "검색·필터 인터랙션 개선 (자동완성·키보드 네비)", "[github]", "이형주", "05/30", "06/04", 6, "할 일", False),
    ("", "", "다크모드 / 반응형 / 모바일 대응 다듬기", "[github]", "이형주", "06/02", "06/05", 4, "할 일", False),
    ("", "", "★ Toss·네이버 증권 수준 UI/UX 품질 목표", "[v0.dev]", "전원", "05/28", "06/05", 9, "핵심", False),

    # ─── 6. Tab3 AI 분석 + 챗봇 (시작) ───
    ("", "[기업개요] Tab3\nAI 분석 (시작)", "AI 종합분석 SWOT 자동 생성 프롬프트 설계", "[드라이브]", "고휘주", "05/30", "06/05", 7, "할 일", False),
    ("", "", "RAG 챗봇 정식 엔진 검토 (Chroma 임베딩 + LangChain)", "[드라이브]", "이형주\n고휘주", "05/30", "06/05", 7, "할 일", False),

    # ─── 7. Sprint 4 마무리 ───
    ("", "2차 배포\nSprint 4 마무리", "데모 리허설 + 발표 자료 정리", "[드라이브]", "전원", "06/04", "06/05", 2, "할 일", False),
    ("", "", "★ 2차 배포 발표 (AWS 클라우드 + 2,695개 종목 + UI 고도화)", "[드라이브]", "전원", "06/05", "06/05", 1, "핵심", False),
    ("", "", "Sprint 4 리뷰 회고", "[드라이브]", "전원", "06/05", "06/05", 1, "할 일", False),
    ("", "4차 멘토링", "4차 멘토링 준비·참석", "[드라이브]", "전원", "06/05", "06/05", 1, "할 일", False),

    # ============ 스프린트 5 ============
    ("스프린트 5", "", "★ 3차 배포 — 메인화면 + AI 챗봇 고도화 + 최종 발표", "", "", "06/08", "06/19", 12, "할 일", True),

    # ─── 8. 메인 화면 신규 ★★★ ───
    ("06/08\n~\n06/19",
     "메인 화면 신규\n(★ 신규 서비스)", "메인화면 기획 — 산업 분류 + 모닝루틴 컨셉 정의", "[드라이브]", "전원", "06/08", "06/10", 3, "할 일", False),
    ("", "", "산업 분류 위젯 (WICS·GICS) — 업종별 시총·등락률 카드", "[github]", "이형주\n유태상", "06/08", "06/14", 7, "할 일", False),
    ("", "", "모닝루틴 위젯 (장 시작 전 체크리스트·뉴스 헤드라인·이슈 종목)", "[github]", "이형주\n고휘주", "06/08", "06/14", 7, "할 일", False),
    ("", "", "★ 산업별 종목 드릴다운 (산업 클릭 → 종목 리스트 → 상세)", "[github]", "이형주", "06/10", "06/16", 7, "핵심", False),
    ("", "", "v0.dev 메인화면 최종 디자인 적용", "[v0.dev]", "이형주", "06/10", "06/16", 7, "할 일", False),

    # ─── 9. Tab3 AI 종합분석·챗봇 고도화 ★★★ ───
    ("", "Tab3 고도화\n(★ AI 핵심 기능)", "AI 종합분석 SWOT 실데이터 연결 + 출처 표기", "[github]", "고휘주", "06/08", "06/15", 8, "할 일", False),
    ("", "", "RAG 챗봇 정식 엔진 구현 (Chroma 20GB 임베딩 활용)", "[github]", "이형주\n고휘주", "06/08", "06/15", 8, "할 일", False),
    ("", "", "종목별 추천 질문 동적 생성 (사전 매핑 → LLM 생성)", "[github]", "고휘주", "06/10", "06/15", 6, "할 일", False),
    ("", "", "답변 출처 표기 + 신뢰도 등급 표시 (사업보고서 페이지)", "[github]", "고휘주", "06/10", "06/15", 6, "할 일", False),
    ("", "", "★ 자유 입력 질문 응답 — OpenAI 비용 모니터링 (잔액 $50)", "[github]", "고휘주\n전원", "06/12", "06/17", 6, "핵심", False),
    ("", "", "챗봇 대화 이력 저장 + 사용자별 컨텍스트 유지", "[github]", "이형주", "06/12", "06/17", 6, "할 일", False),

    # ─── 10. 밸류에이션 2차 완성 ───
    ("", "Valuation\n2차 완성", "4-Way 주식 내재가치 비교 바차트 (DCF·PER·PBR·EV/EBITDA)", "[github]", "유태상", "06/08", "06/15", 8, "할 일", False),
    ("", "", "AI 진단·해설 자동 생성 (LLM 평가/저평가 판단)", "[github]", "유태상", "06/08", "06/15", 8, "할 일", False),
    ("", "", "Audit Trail 자동 생성 (출처 URL·수집 날짜·공식 전체 기록)", "[github]", "유태상", "06/08", "06/15", 8, "할 일", False),
    ("", "", "Excel·PDF 리포트 내보내기 기능", "[github]", "유태상", "06/12", "06/17", 6, "할 일", False),

    # ─── 11. 데이터 정확성 최종 검증 ───
    ("", "전 종목 검증", "★ 2,695개 전 종목 정확성 최종 검증 (자동 비교 스크립트)", "[github]", "이형주\n전원", "06/08", "06/15", 8, "핵심", False),
    ("", "", "신규 상장사 자동 추가 파이프라인 (주1회 갱신)", "[github]", "이형주", "06/12", "06/17", 6, "할 일", False),
    ("", "", "데이터 검증 자동화 (사업보고서 원본 vs DB 일치 체크)", "[github]", "이형주", "06/08", "06/12", 5, "할 일", False),

    # ─── 12. 인프라 최종 고도화 ───
    ("", "인프라 고도화", "API 응답 시간 최적화 (1초 이내 목표) + 캐싱 전략", "[github]", "이형주", "06/08", "06/15", 8, "할 일", False),
    ("", "", "AWS 비용 모니터링 + 안정성 보강 (CloudWatch·AutoScale)", "[aws]", "이형주", "06/10", "06/17", 8, "할 일", False),
    ("", "", "도메인·SSL 정식 설정 (filmn9.com 등)", "[aws]", "이형주", "06/12", "06/18", 7, "할 일", False),
    ("", "", "모바일 반응형 마무리", "[github]", "이형주", "06/15", "06/19", 5, "할 일", False),

    # ─── 13. Sprint 5 마무리 + 최종 발표 ───
    ("", "Sprint 5 마무리\n+ 최종 발표 준비", "3차 배포 (최종 완성본 AWS 배포)", "[aws]", "전원", "06/17", "06/19", 3, "할 일", False),
    ("", "", "Sprint 5 리뷰 발표 (3차 배포 데모)", "[드라이브]", "전원", "06/19", "06/19", 1, "할 일", False),
    ("", "5차 멘토링", "5차 멘토링 준비·참석", "[드라이브]", "전원", "06/16", "06/16", 1, "할 일", False),
    ("", "최종 발표", "최종 발표 자료 작성 (PPT·시연 영상·문서)", "[드라이브]", "전원", "06/16", "06/19", 4, "할 일", False),
    ("", "", "최종 발표 리허설", "[드라이브]", "전원", "06/19", "06/19", 1, "할 일", False),
    ("", "", "🎯 최종 발표 (FILMN9 데모)", "[드라이브]", "전원", "06/20", "06/20", 1, "핵심", False),
]

# ── 입력 ────────────────────────────────────────
for d in data:
    group, task, subtask, link, assignee, start, end, period, status, is_header = d
    if is_header:
        for c in range(1, 10):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill(SPRINT_BG)
            cell.font = Font(name="맑은 고딕", size=11, bold=True, color="1F4E3D")
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.cell(row=row, column=1, value=group)
        ws.cell(row=row, column=3, value=subtask)
        ws.cell(row=row, column=6, value=start)
        ws.cell(row=row, column=7, value=end)
        ws.cell(row=row, column=8, value=period)
        ws.cell(row=row, column=9, value=status)
        ws.row_dimensions[row].height = 26
    else:
        ws.cell(row=row, column=1, value=group)
        ws.cell(row=row, column=2, value=task)
        ws.cell(row=row, column=3, value=subtask)
        ws.cell(row=row, column=4, value=link)
        ws.cell(row=row, column=5, value=assignee)
        ws.cell(row=row, column=6, value=start)
        ws.cell(row=row, column=7, value=end)
        ws.cell(row=row, column=8, value=period)
        ws.cell(row=row, column=9, value=status)

        for c in range(1, 10):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if c in (2, 3):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

        if group:
            ws.cell(row=row, column=1).font = Font(name="맑은 고딕", size=9, bold=True)
            ws.cell(row=row, column=1).fill = fill("F5F5F5")
        if link:
            ws.cell(row=row, column=4).font = Font(name="맑은 고딕", size=9, color="1F4E3D", underline="single")

        a = ac(assignee)
        if a:
            ws.cell(row=row, column=5).fill = fill(a)
            ws.cell(row=row, column=5).font = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")

        ws.cell(row=row, column=6).fill = fill("FFF9C4")
        ws.cell(row=row, column=7).fill = fill("FFF9C4")

        s = sc(status)
        if s:
            ws.cell(row=row, column=9).fill = fill(s)
            ws.cell(row=row, column=9).font = Font(name="맑은 고딕", size=9, bold=True)

        # 핵심 태스크 행 — 살짝 강조
        if status == "핵심":
            for c in range(1, 10):
                ws.cell(row=row, column=c).font = Font(name="맑은 고딕", size=9, bold=True)

        ws.row_dimensions[row].height = 30
    row += 1

# 헤더 행 고정 (목표 박스 + 컬럼헤더 위까지)
ws.freeze_panes = f"A{header_row+1}"

# 저장
out = r"C:\Users\Admin\FILMN9\FILMN9_WBS_v5.xlsx"
wb.save(out)
print(f"저장 완료: {out}")
print(f"총 {row-header_row-1} 행")
